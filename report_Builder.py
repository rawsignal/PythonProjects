import os
import sys

# Set path for DLL
sys.path.append("C:\Sybase15\OCS-15_0\python\python27_64\dll")
import sybpydb

from openpyxl import load_workbook

# Look for import file
base_Path = 'C:\\Users\\lrossign\\Desktop\\Docs\\QA\\SprintManualResults\\116\\'
wb_tmp_Template = 'tmpPrizmIndexupdated.xlsx'
wb_Output_Name_Tmp = 'tmpPrizmIndex.xlsx'
wb_Save_As = raw_input("Please enter a name to save the file as: ")
wb_Output_Name = wb_Save_As + 'prizmIndex.xlsx'
wb_Simmons = 'PRIZM Premier_Simmons Input File_Supermarkets shopped last 30 days - Walmart.xlsx'
wb_Scar = 'ScarboroughPRIZMPremier(68)InputFile.xlsx'

class PRIZM_Data(object) :

    def __init__(self, prizm_Code_List, prizm_Index_List, prizm_Penetration_List, prizm_Base_Percent_List) :
        self.prizm_Code_List = prizm_Code_List
        self.prizm_Index_List = prizm_Index_List
        self.prizm_Penetration_List = prizm_Penetration_List
        self.prizm_Base_Percent_List = prizm_Base_Percent_List

    def get_prizm_Code_List(self) :
        return self.prizm_Code_List

    def get_prizm_Index_List(self) :
        return self.prizm_Index_List

    def get_prizm_Penetration_List(self) :
        return self.prizm_Penetration_List

    def get_prizm_Base_Percent_List(self) :
        return self.prizm_Percent_List

    def set_prizm_Code_List(self, prizm_Code_List) :
        self.prizm_Code_List = prizm_Code_List

    def set_prizm_Index_List(self, prizm_Index_List) :
        self.prizm_Index_List = prizm_Index_List

    def set_prizm_Penetration_List(self, prizm_Penetration_List) :
        self.prizm_Penetration_List = prizm_Penetration_List

    def set_prizm_Base_Percent_List(self, prizm_Base_Percent_List) :
        self.prizm_Base_Percent_List = prizm_Base_Percent_List

class Stored_Excel(object) :

    def __init__(self, col_B, col_C, col_D, col_E) :
        self.col_B = col_B
        self.col_C = col_C
        self.col_D = col_D
        self.col_E = col_E


    def get_col_B(self) :
        return self.col_B

    def get_col_C(self) :
        return self.col_C

    def get_col_D(self) :
        return self.col_D

    def get_col_E(self) :
        return self.col_E

    def set_col_B(self, col_B) :
        self.col_B = col_B

    def set_col_C(self, col_C) :
        self.col_C = col_C

    def set_col_D(self, col_D) :
        self.col_D = col_D

    def set_col_E(self, col_E) :
        self.col_E = col_E

class Scar_Stored_Excel(Stored_Excel) :

    def __init__(self, col_A, col_B, col_C, col_D, col_E, col_F) :
        Stored_Excel.__init__(self, col_B, col_C, col_D, col_E)
        self.col_A = col_A
        self.col_F = col_F

    def get_col_A(self) :
        return self.col_A

    def get_col_F(self) :
        return self.col_F

    def set_col_A(self, col_A) :
        self.col_A = col_A

    def set_col_F(self, col_F) :
        self.col_F = col_F

def imp_Scar_Excel_To_Memory(wb_Name) :
    wb = handle_Workbook(base_Path, wb_Name, True)
    sheet = handle_Sheet(wb)

    col_A = []
    col_B = []
    col_C = []
    col_D = []
    col_E = []
    col_F = []

    for i in range(11, sheet.max_row - 9) :
        col_A.append(sheet.cell(row=i, column=1).value)
        col_B.append(sheet.cell(row=i, column=2).value)
        col_C.append(sheet.cell(row=i, column=3).value)
        col_D.append(sheet.cell(row=i, column=4).value)
        col_E.append(sheet.cell(row=i, column=5).value)
        col_F.append(sheet.cell(row=i, column=6).value)

    new_Excel = Scar_Stored_Excel(col_A, col_B, col_C, col_D, col_E, col_F)

    return new_Excel
     

def imp_Excel_To_Memory(wb_Name) :
    wb = handle_Workbook(base_Path, wb_Name, True)
    sheet = handle_Sheet(wb)

    col_B = []
    col_C = []
    col_D = []
    col_E = []

    for i in range(13, sheet.max_row - 4) :
        col_B.append(sheet.cell(row=i, column=2).value)
        col_C.append(sheet.cell(row=i, column=3).value)
        col_D.append(sheet.cell(row=i, column=4).value)
        col_E.append(sheet.cell(row=i, column=5).value)

    new_Excel = Stored_Excel(col_B, col_C, col_D, col_E)

    return new_Excel


def handle_Workbook(base_Path, file_Name, flag_RO) :
    if flag_RO == False :
        wb = load_workbook(base_Path + file_Name)
    else :
        wb = load_workbook(base_Path + file_Name, read_only=True)

    return wb

def handle_Sheet(wb) :
    sheet_Name = wb.get_sheet_names()
    sheet = wb.get_sheet_by_name(sheet_Name[0])

    return sheet


def export_Index_To_Excel(PRIZM_Data) :
    
    wb_tmp = handle_Workbook(base_Path, wb_tmp_Template, False)
    wb_tmp.save(base_Path + wb_Output_Name)
    wb = handle_Workbook(base_Path, wb_Output_Name, False)
    sheet = handle_Sheet(wb)

    index_Column = 3
    start_Row = 4
    penetration_Column = 4
    base_Percents_Column = 5

    for i in xrange(0, len(PRIZM_Data.prizm_Code_List)) :
        sheet.cell(row=start_Row, column=index_Column).value = PRIZM_Data.prizm_Index_List[i]
        sheet.cell(row=start_Row, column=penetration_Column).value = PRIZM_Data.prizm_Penetration_List[i]
        sheet.cell(row=start_Row, column=base_Percents_Column).value = PRIZM_Data.prizm_Base_Percent_List[i]
        start_Row += 1

    wb.save(base_Path + wb_Output_Name)

def query_Stored_Excel(Stored_Excel) :
    prizm_Code_Name = []
    prizm_Index_List = []
    prizm_Penetration_List = []
    prizm_Base_Percent_List = []

    
    index_Inds = [i for i, x in enumerate(Stored_Excel.col_C) if x == "Index"]
    vert_Inds = [i for i, x in enumerate(Stored_Excel.col_C) if x == "Vertical %"]

    for x in list(index_Inds) :
        prizm_Code_Name.append(Stored_Excel.col_B[x])
        prizm_Index_List.append(Stored_Excel.col_E[x])

    for x in list(vert_Inds) :
        prizm_Base_Percent_List.append(Stored_Excel.col_D[x])
        prizm_Penetration_List.append(Stored_Excel.col_E[x])

    if prizm_Code_Name[0] == 'Total' :
        del prizm_Code_Name[0]
        del prizm_Index_List[0]
        del prizm_Penetration_List[0]
        del prizm_Base_Percent_List[0]
    
    prizm_Data = PRIZM_Data(prizm_Code_Name, prizm_Index_List, prizm_Penetration_List, prizm_Base_Percent_List)
    return prizm_Data

def query_Scar_Stored_Excel(Scar_Stored_Excel) :
    prizm_Code_Name = []
    prizm_Index_List = []
    prizm_Penetration_List = []
    prizm_Base_Percent_List = []

    for prizmname in list(Scar_Stored_Excel.col_A) :
        prizm_Code_Name.append(prizmname)

    for prizmindex in list(Scar_Stored_Excel.col_F) :
        prizm_Index_List.append(prizmindex)

    for prizmbase in list(Scar_Stored_Excel.col_C) :
        prizm_Base_Percent_List.append(prizmbase)

    for prizmpen in list(Scar_Stored_Excel.col_E) :
        prizm_Penetration_List.append(prizmpen)

    prizm_Data = PRIZM_Data(prizm_Code_Name, prizm_Index_List, prizm_Penetration_List, prizm_Base_Percent_List)
    return prizm_Data

    
def main() :
    flag = input("Which file should be processed? 1 for Scar, 2 for Simmons (note, this is for testing the code, this is not a perm solution): ")
    print(flag)
    if flag == 1 :
        test_Excel = imp_Scar_Excel_To_Memory(wb_Scar)
        test = query_Scar_Stored_Excel(test_Excel)
        export_Index_To_Excel(test)
    elif flag == 2 :
        test_Excel = imp_Excel_To_Memory(wb_Simmons)
        test = query_Stored_Excel(test_Excel)
        export_Index_To_Excel(test)

main()

# test the export excel code
#for i in xrange(0, len(test_Excel.col_B)) :
	#print("\n {0} , {1} , {2} , {3} ".format(test_Excel.col_B[i], test_Excel.col_C[i], test_Excel.col_D[i], test_Excel.col_E[i]))
    
        

    
    
